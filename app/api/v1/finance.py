"""
Finance API — Payment, Expense, Salary — /api/v1/finance/
"""
import uuid
from typing import Annotated, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from datetime import datetime, timezone

from app.database import get_db
from app.models.user import User, UserRole
from app.models.finance import Payment, PaymentStatus, Expense, Salary, SalaryStatus
from app.models.student import Student
from app.models.academic import Enrollment
from app.core.dependencies import require_roles, get_current_active_user
from app.schemas import (
    PaymentCreate, PaymentUpdate, PaymentResponse, PaymentBriefResponse,
    ExpenseCreate, ExpenseUpdate, ExpenseResponse,
    SalaryCreate, SalaryUpdate, SalaryPayRequest, SalaryResponse,
    PaginatedResponse, MessageResponse,
)

router = APIRouter(prefix="/finance", tags=["Finance (Moliya)"])

AnyStaff = Depends(require_roles(UserRole.ADMIN, UserRole.MANAGER))
AdminOnly = Depends(require_roles(UserRole.ADMIN))
StaffAndTeacher = Depends(require_roles(UserRole.ADMIN, UserRole.MANAGER, UserRole.TEACHER))


# ═══════════════════════════════════════════════════════════════════════════════
# PAYMENTS
# ═══════════════════════════════════════════════════════════════════════════════
from sqlalchemy.orm import joinedload

@router.get("/payments/", response_model=PaginatedResponse[PaymentBriefResponse], summary="To'lovlar ro'yxati")
async def list_payments(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, AnyStaff],
    search: Optional[str] = Query(None, description="O'quvchi ismi yoki raqami"),
    student_id: Optional[uuid.UUID] = None,
    period_month: Optional[int] = Query(None, ge=1, le=12),
    period_year: Optional[int] = Query(None, ge=2020),
    pay_status: Optional[PaymentStatus] = Query(None, alias="status"),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=10000),
):
    query = select(Payment).join(Student, Payment.student_id == Student.id, isouter=True).options(
        joinedload(Payment.student),
        joinedload(Payment.enrollment).joinedload(Enrollment.group)
    )
    if search:
        query = query.where(
            Student.full_name.ilike(f"%{search}%") | Student.phone.ilike(f"%{search}%")
        )
    if student_id:
        query = query.where(Payment.student_id == student_id)
    if period_month:
        query = query.where(func.extract('month', Payment.created_at) == period_month)
    if period_year:
        query = query.where(func.extract('year', Payment.created_at) == period_year)
    if pay_status:
        query = query.where(Payment.status == pay_status)

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    payments = (await db.execute(query.offset(skip).limit(limit).order_by(Payment.created_at.desc()))).scalars().all()
    return PaginatedResponse.create(data=payments, total=total, skip=skip, limit=limit)


@router.post("/payments/", response_model=PaymentResponse, status_code=201, summary="To'lov qabul qilish")
async def create_payment(
    data: PaymentCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, AnyStaff],
):
    # O'quvchi mavjudligini tekshirish
    student = (await db.execute(select(Student).where(Student.id == data.student_id))).scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="O'quvchi topilmadi")

    payment = Payment(
        **data.model_dump(),
        created_by=current_user.id,
        status=PaymentStatus.CONFIRMED,
    )
    db.add(payment)

    # O'quvchi balansini yangilash
    student.balance += data.amount
    
    # Tizim xabarini yaratish (Barcha adminlarga yuborish)
    from app.models.system import Notification, NotificationType, NotificationChannel, NotificationStatus
    admins = (await db.execute(select(User).where(User.role == UserRole.ADMIN))).scalars().all()
    
    for admin in admins:
        notif = Notification(
            user_id=admin.id,
            title="To'lov qabul qilindi",
            body=f"{student.full_name} dan {data.amount:,.0f} UZS to'lov qabul qilindi.",
            notif_type=NotificationType.SYSTEM_ALERT,
            channel=NotificationChannel.IN_APP,
            status=NotificationStatus.SENT
        )
        db.add(notif)

    await db.flush()
    await db.refresh(payment)
    
    # ── TELEGRAM BOT ORQALI CHEK (RECEIPT) YUBORISH ──
    from app.models.student import Parent
    from aiogram import Bot
    from aiogram.exceptions import TelegramAPIError
    from app.config import settings
    
    if student.parent_id and settings.TELEGRAM_BOT_TOKEN:
        parent = (await db.execute(select(Parent).where(Parent.id == student.parent_id))).scalar_one_or_none()
        if parent and parent.telegram_id:
            try:
                bot = Bot(token=settings.TELEGRAM_BOT_TOKEN)
                months_uz = ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun", "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]
                month_name = months_uz[payment.period_month - 1]
                
                method_uz = {
                    "cash": "Naqd pul",
                    "card": "Plastik karta",
                    "transfer": "Bank o'tkazmasi",
                    "other": "Boshqa"
                }.get(payment.method.value, payment.method.value)
                
                receipt_msg = (
                    f"🧾 <b>To'lov qabul qilindi!</b>\n\n"
                    f"👤 <b>O'quvchi:</b> {student.full_name}\n"
                    f"💰 <b>Miqdor:</b> {payment.amount:,.0f} UZS\n"
                    f"📆 <b>To'lov davri:</b> {month_name}, {payment.period_year}-yil\n"
                    f"💳 <b>To'lov usuli:</b> {method_uz}\n\n"
                    f"<i>Ishonchingiz uchun rahmat! EduCRM</i>"
                )
                
                await bot.send_message(chat_id=parent.telegram_id, text=receipt_msg, parse_mode="HTML")
                await bot.session.close()
            except Exception as e:
                print(f"Telegram yuborishda xatolik: {e}")
                
    return payment


@router.put("/payments/{payment_id}", response_model=PaymentResponse, summary="To'lov holatini o'zgartirish")
async def update_payment(
    payment_id: uuid.UUID,
    data: PaymentUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, AdminOnly],
):
    payment = (await db.execute(select(Payment).where(Payment.id == payment_id))).scalar_one_or_none()
    if not payment:
        raise HTTPException(status_code=404, detail="To'lov topilmadi")

    # CANCELLED bo'lganda balansdan olib tashlash
    if data.status == PaymentStatus.CANCELLED and payment.status == PaymentStatus.CONFIRMED:
        student = (await db.execute(select(Student).where(Student.id == payment.student_id))).scalar_one()
        student.balance = max(0, student.balance - payment.amount)

    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(payment, key, value)

    await db.flush()
    await db.refresh(payment)
    return payment


# ═══════════════════════════════════════════════════════════════════════════════
# EXPENSES
# ═══════════════════════════════════════════════════════════════════════════════
@router.get("/expenses/", response_model=PaginatedResponse[ExpenseResponse], summary="Xarajatlar ro'yxati")
async def list_expenses(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, AnyStaff],
    period_month: Optional[int] = Query(None, ge=1, le=12),
    period_year: Optional[int] = Query(None, ge=2020),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=10000),
):
    query = select(Expense)
    if period_month:
        query = query.where(func.extract('month', Expense.expense_date) == period_month)
    if period_year:
        query = query.where(func.extract('year', Expense.expense_date) == period_year)
        
    query = query.order_by(Expense.expense_date.desc())
    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    expenses = (await db.execute(query.offset(skip).limit(limit))).scalars().all()
    return PaginatedResponse.create(data=expenses, total=total, skip=skip, limit=limit)


@router.post("/expenses/", response_model=ExpenseResponse, status_code=201, summary="Xarajat qo'shish")
async def create_expense(
    data: ExpenseCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, AnyStaff],
):
    expense = Expense(**data.model_dump(), created_by=current_user.id)
    db.add(expense)
    await db.flush()
    await db.refresh(expense)
    return expense


@router.delete("/expenses/{expense_id}", response_model=MessageResponse, summary="Xarajatni o'chirish")
async def delete_expense(
    expense_id: uuid.UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, AdminOnly],
):
    expense = (await db.execute(select(Expense).where(Expense.id == expense_id))).scalar_one_or_none()
    if not expense:
        raise HTTPException(status_code=404, detail="Xarajat topilmadi")
    await db.delete(expense)
    return MessageResponse(detail="Xarajat o'chirildi")


# ═══════════════════════════════════════════════════════════════════════════════
# SALARIES
# ═══════════════════════════════════════════════════════════════════════════════
@router.get("/salaries/", response_model=PaginatedResponse[SalaryResponse], summary="Maoshlar ro'yxati")
async def list_salaries(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, StaffAndTeacher],
    user_id: Optional[uuid.UUID] = None,
    period_month: Optional[int] = Query(None, ge=1, le=12),
    period_year: Optional[int] = Query(None, ge=2020),
    sal_status: Optional[SalaryStatus] = Query(None, alias="status"),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=10000),
):
    # O'qituvchi faqat o'z maoshini ko'radi
    if current_user.role == UserRole.TEACHER:
        user_id = current_user.id

    query = select(Salary).options(joinedload(Salary.user))
    if user_id:
        query = query.where(Salary.user_id == user_id)
    if period_month:
        query = query.where(Salary.period_month == period_month)
    if period_year:
        query = query.where(Salary.period_year == period_year)
    if sal_status:
        query = query.where(Salary.status == sal_status)

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    salaries = (await db.execute(query.offset(skip).limit(limit).order_by(Salary.created_at.desc()))).scalars().all()
    return PaginatedResponse.create(data=salaries, total=total, skip=skip, limit=limit)


@router.get("/salaries/calculate", summary="Maoshni avtomatik hisoblash")
async def calculate_salary(
    user_id: uuid.UUID,
    period_month: int = Query(..., ge=1, le=12),
    period_year: int = Query(..., ge=2020),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    from app.models.academic import Group, Enrollment
    # O'qituvchi bo'lsa faqat o'zini hisoblay oladi
    if current_user.role == UserRole.TEACHER and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Faqat o'z maoshingizni hisoblay olasiz")
        
    query = (
        select(
            func.sum(Payment.amount * (Group.teacher_salary_pct / 100)).label("calculated_salary")
        )
        .select_from(Payment)
        .join(Enrollment, Payment.enrollment_id == Enrollment.id)
        .join(Group, Enrollment.group_id == Group.id)
        .where(
            and_(
                Group.teacher_id == user_id,
                Payment.period_month == period_month,
                Payment.period_year == period_year,
                Payment.status == PaymentStatus.CONFIRMED
            )
        )
    )
    result = (await db.execute(query)).scalar()
    return {"calculated_salary": float(result or 0)}

@router.post("/salaries/", response_model=SalaryResponse, status_code=201, summary="Maosh hisoblash")
async def create_salary(
    data: SalaryCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, AdminOnly],
):
    # Takroriy hisoblashni tekshirish
    existing = (await db.execute(
        select(Salary).where(
            and_(
                Salary.user_id == data.user_id,
                Salary.period_month == data.period_month,
                Salary.period_year == data.period_year,
            )
        )
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"{data.period_month}/{data.period_year} uchun maosh allaqachon hisoblangan",
        )

    salary = Salary(**data.model_dump())
    db.add(salary)
    await db.flush()
    
    # Munosabatlarni (user) yuklash uchun qayta o'qiymiz
    result = await db.execute(
        select(Salary).options(joinedload(Salary.user)).where(Salary.id == salary.id)
    )
    salary = result.scalar_one()
    
    return salary


@router.post("/salaries/{salary_id}/pay", response_model=SalaryResponse, summary="Maosh to'lash")
async def pay_salary(
    salary_id: uuid.UUID,
    data: SalaryPayRequest,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, AdminOnly],
):
    salary = (await db.execute(
        select(Salary).options(joinedload(Salary.user)).where(Salary.id == salary_id)
    )).scalar_one_or_none()
    if not salary:
        raise HTTPException(status_code=404, detail="Maosh topilmadi")
    if salary.status == SalaryStatus.PAID:
        raise HTTPException(status_code=400, detail="Bu maosh allaqachon to'langan")

    salary.status = SalaryStatus.PAID
    salary.paid_at = datetime.now(timezone.utc)
    salary.paid_by = current_user.id
    if data.comment:
        salary.comment = data.comment

    await db.flush()
    await db.refresh(salary)
    return salary
@router.get("/export/excel", summary="Moliya hisobotlarini Excel (XLSX) formatida yuklash")
async def export_excel(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, AnyStaff],
    tab: str = Query(..., description="Qaysi bo'limni yuklash: payments, expenses, salaries"),
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2020),
):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        import io
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl o'rnatilmagan! (pip install openpyxl)")

    wb = openpyxl.Workbook()
    ws = wb.active

    if tab == "payments":
        ws.title = "To'lovlar"
        headers = ["ID", "Sana", "O'quvchi Ismi", "O'quvchi Tel", "Guruh", "Summa", "To'lov Usuli", "Status"]
        ws.append(headers)

        query = select(Payment).options(
            joinedload(Payment.student),
            joinedload(Payment.enrollment).joinedload(Enrollment.group)
        )
        if month: query = query.where(func.extract('month', Payment.created_at) == month)
        if year: query = query.where(func.extract('year', Payment.created_at) == year)
        query = query.order_by(Payment.created_at.desc())
        
        payments = (await db.execute(query)).scalars().all()
        total_income = 0
        for p in payments:
            student_name = p.student.full_name if p.student else "Noma'lum"
            student_phone = p.student.phone if p.student else ""
            group_name = p.group_name or ""
            amount = float(p.amount)
            total_income += amount
            ws.append([
                str(p.id),
                p.created_at.strftime("%Y-%m-%d %H:%M"),
                student_name,
                student_phone,
                group_name,
                amount,
                p.method.value if p.method else "",
                p.status.value if p.status else ""
            ])
        ws.append([])
        ws.append(["", "", "", "", "JAMI TUSHUM:", total_income, "", ""])
            
    elif tab == "expenses":
        ws.title = "Xarajatlar"
        headers = ["ID", "Sana", "Kategoriya", "Izoh", "Summa"]
        ws.append(headers)
        
        query = select(Expense)
        if month: query = query.where(func.extract('month', Expense.expense_date) == month)
        if year: query = query.where(func.extract('year', Expense.expense_date) == year)
        query = query.order_by(Expense.expense_date.desc())
        
        expenses = (await db.execute(query)).scalars().all()
        cat_map = {
            "rent": "Ijara to'lovi", "salary": "Maosh", "utility": "Kommunal xizmatlar",
            "equipment": "Jihozlar", "marketing": "Marketing va Reklama", "other": "Boshqa xarajatlar"
        }
        total_expense = 0
        for e in expenses:
            cat_name = cat_map.get(e.category.value, e.category.value) if e.category else ""
            amount = float(e.amount)
            total_expense += amount
            ws.append([
                str(e.id),
                e.expense_date.strftime("%Y-%m-%d") if e.expense_date else "",
                cat_name,
                e.description or "",
                amount
            ])
        ws.append([])
        ws.append(["", "", "", "JAMI XARAJAT:", total_expense])
            
    elif tab == "salaries":
        ws.title = "Maoshlar"
        headers = ["ID", "O'qituvchi", "Davr (Oy-Yil)", "Asosiy Maosh", "Bonus (+)", "Jarima (-)", "Jami To'langan", "Status"]
        ws.append(headers)
        
        query = select(Salary).options(joinedload(Salary.user))
        if month: query = query.where(Salary.period_month == month)
        if year: query = query.where(Salary.period_year == year)
        query = query.order_by(Salary.created_at.desc())
        
        salaries = (await db.execute(query)).scalars().all()
        total_salary = 0
        for s in salaries:
            teacher_name = s.user.full_name if s.user else "Noma'lum"
            total = float(s.base_amount + s.bonus_amount - s.penalty_amount)
            total_salary += total
            ws.append([
                str(s.id),
                teacher_name,
                f"{s.period_month}-{s.period_year}",
                float(s.base_amount),
                float(s.bonus_amount),
                float(s.penalty_amount),
                total,
                s.status.value if s.status else ""
            ])
        ws.append([])
        ws.append(["", "", "", "", "", "JAMI MAOSH:", total_salary, ""])
    else:
        raise HTTPException(status_code=400, detail="Noto'g'ri bo'lim nomi (tab)")

    # ── Umumiy Hisobot (Summary) varag'ini qo'shish ──
    # O'sha oydagi barcha tushum va xarajatlarni hisoblash
    ws2 = wb.create_sheet(title="Umumiy Hisobot")
    ws2.append(["Ko'rsatkich", "Summa (UZS)"])
    
    # Jami tushumni hisoblash
    q_pay = select(func.sum(Payment.amount)).where(Payment.status == PaymentStatus.CONFIRMED)
    if month: q_pay = q_pay.where(func.extract('month', Payment.created_at) == month)
    if year: q_pay = q_pay.where(func.extract('year', Payment.created_at) == year)
    total_income_all = (await db.execute(q_pay)).scalar() or 0
    
    # Jami xarajatni hisoblash
    q_exp = select(func.sum(Expense.amount))
    if month: q_exp = q_exp.where(func.extract('month', Expense.expense_date) == month)
    if year: q_exp = q_exp.where(func.extract('year', Expense.expense_date) == year)
    total_expense_all = (await db.execute(q_exp)).scalar() or 0
    
    # Jami maoshlarni hisoblash (Faqat to'langanlari)
    q_sal = select(func.sum(Salary.base_amount + Salary.bonus_amount - Salary.penalty_amount)).where(Salary.status == SalaryStatus.PAID)
    if month: q_sal = q_sal.where(Salary.period_month == month)
    if year: q_sal = q_sal.where(Salary.period_year == year)
    total_salary_all = (await db.execute(q_sal)).scalar() or 0
    
    net_profit = float(total_income_all) - float(total_expense_all) - float(total_salary_all)
    
    ws2.append(["Jami Tushum (To'lovlar):", float(total_income_all)])
    ws2.append(["Jami Xarajatlar:", float(total_expense_all)])
    ws2.append(["Jami Maoshlar (To'langan):", float(total_salary_all)])
    ws2.append([])
    ws2.append(["QOLDIQ (Sof Foyda/Zarar):", net_profit])
    
    ws2.append([])
    ws2.append([])
    ws2.append(["--- XARAJATLAR JADVALI ---"])
    ws2.append(["Sana", "Kategoriya", "Izoh", "Summa (UZS)"])
    
    q_exp_list = select(Expense)
    if month: q_exp_list = q_exp_list.where(func.extract('month', Expense.expense_date) == month)
    if year: q_exp_list = q_exp_list.where(func.extract('year', Expense.expense_date) == year)
    q_exp_list = q_exp_list.order_by(Expense.expense_date.desc())
    all_expenses = (await db.execute(q_exp_list)).scalars().all()
    
    cat_map = {
        "rent": "Ijara to'lovi", "salary": "Maosh", "utility": "Kommunal xizmatlar",
        "equipment": "Jihozlar", "marketing": "Marketing va Reklama", "other": "Boshqa xarajatlar"
    }
    for e in all_expenses:
        cat_name = cat_map.get(e.category.value, e.category.value) if e.category else ""
        ws2.append([
            e.expense_date.strftime("%Y-%m-%d") if e.expense_date else "",
            cat_name,
            e.description or "",
            float(e.amount)
        ])
        
    ws2.append([])
    ws2.append([])
    ws2.append(["--- MAOSHLAR JADVALI ---"])
    ws2.append(["O'qituvchi", "Davr", "Asosiy", "Bonus", "Jarima", "Jami To'langan", "Status"])
    
    q_sal_list = select(Salary).options(joinedload(Salary.user))
    if month: q_sal_list = q_sal_list.where(Salary.period_month == month)
    if year: q_sal_list = q_sal_list.where(Salary.period_year == year)
    q_sal_list = q_sal_list.order_by(Salary.created_at.desc())
    all_salaries = (await db.execute(q_sal_list)).scalars().all()
    
    for s in all_salaries:
        teacher_name = s.user.full_name if s.user else "Noma'lum"
        total = float(s.base_amount + s.bonus_amount - s.penalty_amount)
        ws2.append([
            teacher_name,
            f"{s.period_month}-{s.period_year}",
            float(s.base_amount),
            float(s.bonus_amount),
            float(s.penalty_amount),
            total,
            s.status.value if s.status else ""
        ])

    # Ustunlar kengligini sozlash (ws2 uchun)
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws2.column_dimensions[column].width = max_length + 2

    # Ustunlar kengligini sozlash (ws uchun)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Faylga yozish
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=finance_{tab}_{year}_{month}.xlsx"}
    )
